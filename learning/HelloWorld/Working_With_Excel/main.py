import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')

sheet = wb.active

total = 0
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    total += int(cell)

sheet['F4'].value = total
    
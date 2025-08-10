import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb = xl.load_workbook('revision.xlsx')

sheet = wb.active

for row in range(2, 8):
    quantity = sheet.cell(row=row, column=2).value
    price = sheet.cell(row, 3).value
    total = quantity * price
    total_cell = sheet.cell(row, 4)
    total_cell.value = total

prices = 0
for row in range(2, 8):
    prices += sheet.cell(row, 4).value
    sheet['D10'].value = prices

values = Reference(sheet, min_row=2,
                   max_row=7,
                   min_col=2,
                   max_col=2)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'F2')

wb.save('revision.xlsx')
